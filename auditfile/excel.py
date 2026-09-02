"""Excel-export van de analyse.

De tabbladnamen worden opgebouwd uit de boekjaren in de auditfiles zelf, zodat
er geen jaartallen in de code staan die na een jaarwisseling niet meer kloppen.
"""
from __future__ import annotations

from io import BytesIO

import pandas as pd

from . import controls, vat
from .capability import build_bestandsprofiel, build_relatiedekking
from .comparison import (
    build_jaarovergang,
    build_jaarovergang_verloop,
    build_rubriek_vergelijking,
    controleer_bestandenpaar,
)
from .findings import Materialiteit, grondslag_omzet, pas_review_toe, verzamel_bevindingen
from .integrity import controleer_auditfile
from .model import Auditfile
from .relatiesaldi import build_relatiesaldi, build_relatiesaldo_aansluiting

# Kolomnamen waarvan de inhoud als bedrag moet worden opgemaakt.
BEDRAG_FRAGMENTEN = (
    "bedrag",
    "saldo",
    "mutatie",
    "grondslag",
    "btw",
    "totaal",
    "verschil",
    "omzet",
    "loonkosten",
    "afwijking",
    "gemiddeld",
    "gefactureerd",
    "afgewikkeld",
)
PERCENTAGE_FRAGMENTEN = ("pct", "percentage", "aandeel", "tarief")

# Nederlandse getalnotatie met een rood minteken bij negatieve bedragen.
BEDRAGNOTATIE = '#,##0.00;[Red]-#,##0.00'
PERCENTAGENOTATIE = '0.0"%"'
DATUMNOTATIE = "DD-MM-YYYY"

MAXIMALE_TABBLADNAAM = 31


def _is_bedrag(kolom: str) -> bool:
    genormaliseerd = str(kolom).lower().replace("-", "_")
    if any(fragment in genormaliseerd for fragment in PERCENTAGE_FRAGMENTEN):
        return False
    return any(fragment in genormaliseerd for fragment in BEDRAG_FRAGMENTEN)


def _is_percentage(kolom: str) -> bool:
    genormaliseerd = str(kolom).lower().replace("-", "_")
    return any(fragment in genormaliseerd for fragment in PERCENTAGE_FRAGMENTEN)


def _veilige_tabbladnaam(naam: str, bestaand: set[str]) -> str:
    """Excel staat maximaal 31 tekens toe en geen dubbele namen."""
    schoon = "".join(teken for teken in naam if teken not in r"[]:*?/\\")[:MAXIMALE_TABBLADNAAM]
    kandidaat = schoon
    teller = 2
    while kandidaat in bestaand:
        achtervoegsel = f" {teller}"
        kandidaat = schoon[: MAXIMALE_TABBLADNAAM - len(achtervoegsel)] + achtervoegsel
        teller += 1
    bestaand.add(kandidaat)
    return kandidaat


def _leeg_of_data(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame({"Melding": ["Geen gegevens gevonden"]})
    return df.copy()


def _neutraliseer_formules(worksheet) -> None:
    """Voorkom dat tekst uit het auditfile een werkende formule wordt.

    Openpyxl leidt het celtype af uit de waarde: een tekst die met "=" begint
    wordt als formule weggeschreven. Een rekeningomschrijving, relatienaam of
    ondernemingsnaam uit een auditfile is invoer van buiten en mag in de export
    geen actieve inhoud worden. De waarde blijft ongewijzigd; alleen het
    celtype wordt op tekst vastgezet, zodat de export laat zien wat er in het
    bestand staat.
    """
    for rij in worksheet.iter_rows():
        for cel in rij:
            if isinstance(cel.value, str) and cel.value.startswith("="):
                cel.data_type = "s"


def _opmaak(worksheet, df: pd.DataFrame) -> None:
    """Zet filter, vaste kop, kolombreedtes en getalnotatie."""
    from openpyxl.utils import get_column_letter

    worksheet.freeze_panes = "A2"
    if worksheet.max_row >= 1:
        worksheet.auto_filter.ref = worksheet.dimensions

    for nummer, kolom in enumerate(df.columns, start=1):
        letter = get_column_letter(nummer)
        breedte = len(str(kolom))
        for waarde in df[kolom].head(500):
            if pd.notna(waarde):
                breedte = max(breedte, len(str(waarde)))
        worksheet.column_dimensions[letter].width = min(max(breedte + 2, 12), 60)

        notatie = None
        if pd.api.types.is_datetime64_any_dtype(df[kolom]):
            notatie = DATUMNOTATIE
        elif pd.api.types.is_numeric_dtype(df[kolom]):
            if _is_percentage(kolom):
                notatie = PERCENTAGENOTATIE
            elif _is_bedrag(kolom):
                notatie = BEDRAGNOTATIE
        if notatie:
            for cel in worksheet[letter][1:]:
                cel.number_format = notatie


def _jaarlabel(af: Auditfile) -> str:
    return af.boekjaar or af.bestandsnaam or "onbekend"


def bouw_werkbladen(
    huidig: Auditfile,
    vorig: Auditfile,
    vergelijking: pd.DataFrame,
    btw_mapping: dict[str, str] | None = None,
    aangifte: dict[str, float] | None = None,
    aftrekbaarheid: dict[str, float] | None = None,
    grondslagen: dict[str, float] | None = None,
    materialiteit: Materialiteit | None = None,
    review: dict | None = None,
) -> dict[str, pd.DataFrame]:
    """Stel alle werkbladen samen die in de export komen."""
    jaar_huidig = _jaarlabel(huidig)
    jaar_vorig = _jaarlabel(vorig)

    gebruik = vat.pas_mapping_toe(vat.build_vat_usage(huidig), btw_mapping, aftrekbaarheid)
    rubrieken = vat.build_rubric_summary(gebruik, aangifte, grondslagen)
    materialiteit = materialiteit or Materialiteit(grondslag=grondslag_omzet(huidig))
    bevindingen = pas_review_toe(
        verzamel_bevindingen(
            huidig,
            vorig,
            gebruik=gebruik,
            vergelijking=vergelijking,
            aangifte=aangifte,
            grondslagen=grondslagen,
            materialiteit=materialiteit,
        ),
        review,
    )

    saldo = controls.voeg_rgs_rubriek_toe(huidig.saldo)
    is_balans = saldo["accTp"].astype(str).str.upper().eq("B")

    mutatie_kolommen = [
        "datum",
        "periode",
        "tx_jrnID",
        "tx_jrn_desc",
        "tx_nr",
        "tx_desc",
        "line_nr",
        "line_accID",
        "accDesc",
        "accTp",
        "RGScode",
        "line_docRef",
        "line_invRef",
        "line_custSupID",
        "line_desc",
        "bedrag",
        "vat_vatID",
        "vat_vatPerc",
        "btw_bedrag",
    ]
    mutaties = huidig.lines[[k for k in mutatie_kolommen if k in huidig.lines.columns]].copy()

    return {
        "Bedrijfsgegevens": huidig.company_info_frame(),
        "Bevindingen": bevindingen,
        "Bestandsgegevens": build_bestandsprofiel(huidig),
        "Relatiedekking": build_relatiedekking(huidig),
        "Bestandenpaar": controleer_bestandenpaar(vorig, huidig),
        f"Integriteit {jaar_huidig}": controleer_auditfile(huidig),
        f"Integriteit {jaar_vorig}": controleer_auditfile(vorig),
        "Jaarovergang": build_jaarovergang_verloop(vorig, huidig),
        "Jaarovergang per rekening": build_jaarovergang(vorig, huidig),
        f"Grootboek {jaar_huidig}": saldo,
        f"Balans {jaar_huidig}": saldo[is_balans],
        f"Resultaat {jaar_huidig}": saldo[~is_balans],
        "Mutaties": mutaties.sort_values(["line_accID", "datum", "tx_nr"], na_position="last"),
        f"Vergelijking {jaar_vorig}-{jaar_huidig}": vergelijking,
        "Vergelijking per rubriek": build_rubriek_vergelijking(vergelijking),
        "Btw-codetabel": huidig.vat_codes,
        "Btw per code": gebruik,
        "Btw per rubriek": rubrieken,
        "Btw-rondrekening": vat.build_vat_ledger_flow(huidig, rubrieken),
        "Btw-signalen": vat.build_vat_anomalies(huidig, gebruik),
        "Btw-drilldown": vat.build_vat_drilldown(huidig),
        "Periodieke controles": controls.build_periodieke_controles(huidig),
        "Ongebruikelijke boekingen": controls.build_ongebruikelijke_boekingen(huidig),
        "Balanssignalen": controls.build_balanspost_signalen(huidig),
        "Fiscale signalen": controls.build_fiscale_signalen(huidig),
        "Omzet per periode": controls.build_omzet_per_periode(huidig),
        "Loonkosten per periode": controls.build_personeelskosten_per_periode(huidig),
        "Debiteuren gefactureerd": controls.build_relatie_analyse(huidig, "debiteur", top=100),
        "Crediteuren gefactureerd": controls.build_relatie_analyse(huidig, "crediteur", top=100),
        # Alleen gevuld bij XAF 4.0 met openstaande bedragen per relatie; bij een
        # leeg blad zou de export suggereren dat er niets openstaat.
        "Aansluiting relatiesaldi": build_relatiesaldo_aansluiting(huidig),
        "Openstaand per relatie": build_relatiesaldi(huidig),
    }


def build_excel_export(
    huidig: Auditfile,
    vorig: Auditfile,
    vergelijking: pd.DataFrame,
    btw_mapping: dict[str, str] | None = None,
    aangifte: dict[str, float] | None = None,
    aftrekbaarheid: dict[str, float] | None = None,
    grondslagen: dict[str, float] | None = None,
    materialiteit: Materialiteit | None = None,
    review: dict | None = None,
) -> bytes:
    """Bouw het Excelbestand met alle werkbladen."""
    werkbladen = bouw_werkbladen(
        huidig,
        vorig,
        vergelijking,
        btw_mapping,
        aangifte,
        aftrekbaarheid,
        grondslagen,
        materialiteit,
        review,
    )
    uitvoer = BytesIO()
    gebruikte_namen: set[str] = set()

    with pd.ExcelWriter(uitvoer, engine="openpyxl") as writer:
        for naam, blad in werkbladen.items():
            data = _leeg_of_data(blad)
            tabblad = _veilige_tabbladnaam(naam, gebruikte_namen)
            data.to_excel(writer, sheet_name=tabblad, index=False)
            _opmaak(writer.sheets[tabblad], data)
            _neutraliseer_formules(writer.sheets[tabblad])

    return uitvoer.getvalue()


def exportnaam(huidig: Auditfile, vorig: Auditfile) -> str:
    """Bestandsnaam voor de download, zonder klantnaam."""
    return f"auditfile-analyse_{_jaarlabel(vorig)}_{_jaarlabel(huidig)}.xlsx"
