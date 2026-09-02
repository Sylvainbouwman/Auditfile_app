"""Tests op de vergelijking en de Excel-export."""
from __future__ import annotations

from io import BytesIO

import openpyxl
import pandas as pd

from auditfile.comparison import (
    build_opvallende_verschillen,
    build_rubriek_vergelijking,
    compare_saldi,
)
from auditfile.excel import build_excel_export, bouw_werkbladen, exportnaam
from auditfile.parsing import parse_auditfile
from auditfile.demo import Account, AuditfileSpec, Journal, Line, Transaction, build_xaf


# --- Vergelijking -----------------------------------------------------------


def test_vergelijking_van_gelijke_jaren_geeft_geen_verschillen(af_40):
    vergelijking = compare_saldi(af_40, af_40)
    assert (vergelijking["verschil_bedrag"].abs() < 0.005).all()
    assert set(vergelijking["status"]) == {"bestaand"}


def test_nieuwe_en_vervallen_rekeningen():
    basis = AuditfileSpec(
        accounts=[Account("1000", "Kas", "B"), Account("8000", "Omzet", "P")],
        journals=[
            Journal(
                "MEM",
                "Memoriaal",
                [
                    Transaction(
                        "M1",
                        "2025-01-01",
                        1,
                        [Line("1000", "100.00", "D"), Line("8000", "100.00", "C")],
                    )
                ],
            )
        ],
    )
    nieuw = AuditfileSpec(
        accounts=[Account("1000", "Kas", "B"), Account("8100", "Andere omzet", "P")],
        journals=[
            Journal(
                "MEM",
                "Memoriaal",
                [
                    Transaction(
                        "M1",
                        "2025-01-01",
                        1,
                        [Line("1000", "100.00", "D"), Line("8100", "100.00", "C")],
                    )
                ],
            )
        ],
    )
    vergelijking = compare_saldi(
        parse_auditfile("vorig.xaf", build_xaf(basis)),
        parse_auditfile("huidig.xaf", build_xaf(nieuw)),
    ).set_index("rekening")
    assert vergelijking.loc["8100", "status"] == "nieuw"
    assert vergelijking.loc["8000", "status"] == "vervallen"


def test_verschilpercentage_blijft_leeg_bij_beginstand_nul(af_40):
    leeg = AuditfileSpec(accounts=[Account("8000", "Omzet", "P")], journals=[])
    vergelijking = compare_saldi(parse_auditfile("leeg.xaf", build_xaf(leeg)), af_40)
    nieuw = vergelijking[vergelijking["status"] == "nieuw"]
    assert nieuw["verschil_pct"].isna().all()


def test_bedragen_blijven_numeriek(af_40):
    """De vergelijking mag geen opgemaakte tekst opleveren; dan is sorteren stuk."""
    vergelijking = compare_saldi(af_40, af_40)
    for kolom in ["saldo_vorig", "saldo_huidig", "verschil_bedrag", "verschil_pct"]:
        assert pd.api.types.is_numeric_dtype(vergelijking[kolom]), kolom


def test_rgs_rubriek_wordt_toegevoegd(af_40):
    vergelijking = compare_saldi(af_40, af_40)
    assert "Netto-omzet" in set(vergelijking["RGS-rubriek"])


def test_vergelijking_per_rubriek(af_40):
    vergelijking = compare_saldi(af_40, af_40)
    per_rubriek = build_rubriek_vergelijking(vergelijking)
    assert not per_rubriek.empty
    assert (per_rubriek["verschil_bedrag"].abs() < 0.005).all()


def test_opvallende_verschillen_filtert_op_bedrag_en_percentage(af_40):
    leeg = AuditfileSpec(accounts=[Account("8000", "Omzet", "P")], journals=[])
    vergelijking = compare_saldi(parse_auditfile("leeg.xaf", build_xaf(leeg)), af_40)
    opvallend = build_opvallende_verschillen(vergelijking, minimaal_bedrag=1000.0)
    assert (opvallend["verschil_bedrag"].abs() >= 1000.0).all()


# --- Excel ------------------------------------------------------------------


def test_export_bevat_alle_werkbladen(af_40):
    vergelijking = compare_saldi(af_40, af_40)
    inhoud = build_excel_export(af_40, af_40, vergelijking)
    werkboek = openpyxl.load_workbook(BytesIO(inhoud))
    verwacht = bouw_werkbladen(af_40, af_40, vergelijking)
    assert len(werkboek.sheetnames) == len(verwacht)


def test_tabbladnamen_bevatten_het_boekjaar_uit_het_bestand(af_40):
    """Geen jaartallen in de code: die komen uit de auditfile zelf."""
    vergelijking = compare_saldi(af_40, af_40)
    werkboek = openpyxl.load_workbook(BytesIO(build_excel_export(af_40, af_40, vergelijking)))
    assert "Balans 2025" in werkboek.sheetnames
    assert "Vergelijking 2025-2025" in werkboek.sheetnames


def test_tabbladnamen_blijven_binnen_de_grens_van_excel(af_40):
    vergelijking = compare_saldi(af_40, af_40)
    werkboek = openpyxl.load_workbook(BytesIO(build_excel_export(af_40, af_40, vergelijking)))
    for naam in werkboek.sheetnames:
        assert len(naam) <= 31, naam
        assert not set(naam) & set(r"[]:*?/\\")


def test_bedragen_worden_als_getal_weggeschreven(af_40):
    """Een bedrag moet in Excel optelbaar zijn, niet als tekst staan."""
    vergelijking = compare_saldi(af_40, af_40)
    werkboek = openpyxl.load_workbook(BytesIO(build_excel_export(af_40, af_40, vergelijking)))
    blad = werkboek["Balans 2025"]
    koppen = [cel.value for cel in blad[1]]
    kolom = koppen.index("eindsaldo") + 1
    waarden = [blad.cell(row=rij, column=kolom).value for rij in range(2, blad.max_row + 1)]
    assert any(isinstance(waarde, (int, float)) for waarde in waarden)


def test_bedragkolom_krijgt_nederlandse_notatie(af_40):
    vergelijking = compare_saldi(af_40, af_40)
    werkboek = openpyxl.load_workbook(BytesIO(build_excel_export(af_40, af_40, vergelijking)))
    blad = werkboek["Balans 2025"]
    koppen = [cel.value for cel in blad[1]]
    kolom = koppen.index("eindsaldo") + 1
    assert blad.cell(row=2, column=kolom).number_format == "#,##0.00;[Red]-#,##0.00"


def test_export_werkt_op_een_leeg_auditfile():
    leeg = parse_auditfile("leeg.xaf", build_xaf(AuditfileSpec(accounts=[], journals=[])))
    vergelijking = compare_saldi(leeg, leeg)
    inhoud = build_excel_export(leeg, leeg, vergelijking)
    werkboek = openpyxl.load_workbook(BytesIO(inhoud))
    assert len(werkboek.sheetnames) > 0


def test_exportnaam_bevat_geen_klantnaam(af_40):
    naam = exportnaam(af_40, af_40)
    assert af_40.bedrijfsnaam not in naam
    assert naam.endswith(".xlsx")
